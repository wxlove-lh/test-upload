import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


def generate_excel(transactions, summary=None):
    """
    生成Excel文件
    transactions: Transaction对象列表
    summary: 汇总数据字典（可选）
    返回：BytesIO对象
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "账目明细"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写表头
    headers = ["日期", "类型", "金额", "分类", "供应商", "备注", "状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写数据行
    for row_idx, t in enumerate(transactions, 2):
        ws.cell(row=row_idx, column=1, value=str(t.transaction_date))
        ws.cell(row=row_idx, column=2, value="收入" if t.type == "income" else "支出")
        amount_cell = ws.cell(row=row_idx, column=3, value=float(t.amount))
        amount_cell.number_format = '#,##0.00'
        # 收入绿色，支出红色
        if t.type == "income":
            amount_cell.font = Font(color="008000")
        else:
            amount_cell.font = Font(color="FF0000")
        ws.cell(row=row_idx, column=4, value=t.category)
        ws.cell(row=row_idx, column=5, value=t.supplier)
        ws.cell(row=row_idx, column=6, value=t.notes)
        ws.cell(row=row_idx, column=7, value="已确认" if t.status == "confirmed" else "已修改" if t.status == "modified" else "待确认")

    # 如果有汇总数据，在数据下方添加汇总行
    if summary:
        summary_row = len(transactions) + 3
        ws.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=f"总收入: ¥{summary['total_income']:,.2f}")
        ws.cell(row=summary_row, column=3, value=f"总支出: ¥{summary['total_expense']:,.2f}")
        ws.cell(row=summary_row, column=4, value=f"净利润: ¥{summary['total_profit']:,.2f}")

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 30)

    # 保存到BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
